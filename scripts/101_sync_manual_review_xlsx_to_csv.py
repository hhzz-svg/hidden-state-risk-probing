# -*- coding: utf-8 -*-
"""Sync filled human-review columns from XLSX workbooks back to CSV files."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

JOBS = [
    {
        "name": "experiment4_round2",
        "xlsx": ROOT / "reports/experiment4/manual_review/experiment4_human_check_round2_80.xlsx",
        "csv": ROOT / "reports/experiment4/manual_review/experiment4_human_check_round2_80.csv",
        "sheet": "Experiment4_Round2",
        "key": "human_check_id",
        "sync_cols": ["human_behavior_label", "human_notes", "human_reviewer"],
    },
    {
        "name": "pk_ck_v2_sample44",
        "xlsx": ROOT / "reports/experiment7_optionalA/pk_ck_v2_manual_review_sample44.xlsx",
        "csv": ROOT / "reports/experiment7_optionalA/pk_ck_v2_manual_review_sample44.csv",
        "sheet": "PKCK_v2_Review",
        "key": "pkck_review_id",
        "sync_cols": ["human_pkck_label", "human_notes", "human_reviewer"],
    },
]


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv_rows(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def read_xlsx_values(path: Path, sheet_name: str, key: str, sync_cols: list[str]) -> dict[str, dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"{path} missing sheet {sheet_name!r}")
    ws = wb[sheet_name]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}
    required = [key, *sync_cols]
    missing = [name for name in required if name not in header_index]
    if missing:
        raise KeyError(f"{path} missing columns: {missing}")

    values: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_key = str(row[header_index[key]] or "").strip()
        if not row_key:
            continue
        values[row_key] = {
            col: str(row[header_index[col]] or "").strip()
            for col in sync_cols
        }
    return values


def sync_one(job: dict[str, object]) -> None:
    csv_path = Path(job["csv"])
    rows = read_csv_rows(csv_path)
    if not rows:
        raise ValueError(f"empty CSV: {csv_path}")
    fieldnames = list(rows[0].keys())
    key = str(job["key"])
    sync_cols = list(job["sync_cols"])
    xlsx_values = read_xlsx_values(Path(job["xlsx"]), str(job["sheet"]), key, sync_cols)

    updated_cells = 0
    filled_rows = 0
    for row in rows:
        row_key = str(row.get(key, "")).strip()
        if row_key not in xlsx_values:
            continue
        has_filled = False
        for col in sync_cols:
            new_value = xlsx_values[row_key].get(col, "")
            if row.get(col, "") != new_value:
                row[col] = new_value
                updated_cells += 1
            if new_value:
                has_filled = True
        if has_filled:
            filled_rows += 1

    write_csv_rows(csv_path, rows, fieldnames)
    print(f"{job['name']}: synced {updated_cells} cells; filled rows now {filled_rows}; saved {csv_path}")


def main() -> None:
    for job in JOBS:
        sync_one(job)


if __name__ == "__main__":
    main()
